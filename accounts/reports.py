"""
Excel report generation for Admin Reports section.
Uses openpyxl. Only for admin users.
"""
from django.db.models import Count, Sum, Avg, Q, F, Max, Min
from django.db.models.functions import TruncMonth
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from django.contrib.auth.models import User
from accounts.models import UserProfile
from courses.models import Course, Lesson, Enrollment, LessonCompletion
from exams.models import Exam, ExamResult


def _style_header(ws):
    """Bold header row and auto-adjust column widths."""
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for col in range(1, ws.max_column + 1):
        max_len = 12
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[get_column_letter(col)].width = min(50, max_len)


# -------------------------------------------------------------------
# Student-related reports
# -------------------------------------------------------------------

def build_student_enrollment_report():
    """
    1) Student Enrollment Report
    Columns: Student Name, Course Name, Enrollment Date, Status
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Enrollments"
    headers = ["Student Name", "Course Name", "Enrollment Date", "Status"]
    ws.append(headers)

    enrollments = Enrollment.objects.select_related("student", "course").order_by(
        "student__username", "course__title"
    )
    for e in enrollments:
        status = "Completed" if e.completed else "In Progress"
        ws.append(
            [
                e.student.get_full_name() or e.student.username,
                e.course.title,
                e.enrolled_at.strftime("%Y-%m-%d %H:%M"),
                status,
            ]
        )

    _style_header(ws)
    return wb


def build_course_student_list_report():
    """
    2) Course-wise Student List
    Columns: Course Name, Total Students, List of Enrolled Students
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Course Students"
    headers = ["Course Name", "Total Students", "Enrolled Students"]
    ws.append(headers)

    courses = (
        Course.objects.annotate(total_students=Count("enrollments", distinct=True))
        .prefetch_related("enrollments__student")
        .order_by("title")
    )

    for c in courses:
        students = [
            e.student.get_full_name() or e.student.username for e in c.enrollments.all()
        ]
        ws.append(
            [
                c.title,
                c.total_students,
                ", ".join(students),
            ]
        )

    _style_header(ws)
    return wb


def build_student_progress_report():
    """
    3) Student Progress Report
    Columns: Student Name, Course, Lessons Completed, Exams Attempted, Overall Progress (%)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Progress"
    headers = [
        "Student Name",
        "Course",
        "Lessons Completed",
        "Total Lessons",
        "Exams Attempted",
        "Overall Progress (%)",
    ]
    ws.append(headers)

    # Pre-compute lesson counts per course
    lesson_counts = dict(
        Lesson.objects.values("course_id").annotate(c=Count("id")).values_list(
            "course_id", "c"
        )
    )

    # Pre-compute completed lessons per (student, course)
    completed_map = {}
    for row in (
        LessonCompletion.objects.values("student_id", "lesson__course_id")
        .annotate(c=Count("id"))
        .values("student_id", "lesson__course_id", "c")
    ):
        completed_map[(row["student_id"], row["lesson__course_id"])] = row["c"]

    # Pre-compute exam attempts per (student, course)
    attempts_map = {}
    for row in (
        ExamResult.objects.filter(is_completed=True)
        .values("student_id", "exam__course_id")
        .annotate(c=Count("id"))
        .values("student_id", "exam__course_id", "c")
    ):
        attempts_map[(row["student_id"], row["exam__course_id"])] = row["c"]

    enrollments = Enrollment.objects.select_related("student", "course").order_by(
        "student__username", "course__title"
    )

    for e in enrollments:
        total_lessons = lesson_counts.get(e.course_id, 0) or 0
        lessons_completed = completed_map.get((e.student_id, e.course_id), 0)
        exams_attempted = attempts_map.get((e.student_id, e.course_id), 0)

        if total_lessons > 0:
            progress = round((lessons_completed / total_lessons) * 100, 2)
        else:
            progress = 0.0

        ws.append(
            [
                e.student.get_full_name() or e.student.username,
                e.course.title,
                lessons_completed,
                total_lessons,
                exams_attempted,
                progress,
            ]
        )

    _style_header(ws)
    return wb


def build_student_exam_report():
    """
    4) Student Exam Performance Report
    Columns: Student Name, Course, Exam, Marks Obtained, Total Marks, Percentage
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Exam Performance"
    headers = [
        "Student Name",
        "Course",
        "Exam",
        "Marks Obtained",
        "Total Marks",
        "Percentage",
    ]
    ws.append(headers)

    qs = (
        ExamResult.objects.filter(is_completed=True)
        .select_related("student", "exam", "exam__course")
        .order_by("student__username", "exam__course__title", "exam__title")
    )
    for r in qs:
        ws.append(
            [
                r.student.get_full_name() or r.student.username,
                r.exam.course.title,
                r.exam.title,
                r.score,
                r.total_marks,
                round(r.percentage, 2),
            ]
        )

    _style_header(ws)
    return wb


# -------------------------------------------------------------------
# Teacher-related reports
# -------------------------------------------------------------------

def build_teacher_activity_report():
    """
    5) Teacher Activity Report
    Columns: Teacher Name, Total Courses, Total Lessons, Total Exams
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Teacher Activity"
    headers = ["Teacher Name", "Total Courses", "Total Lessons", "Total Exams"]
    ws.append(headers)

    teachers = User.objects.filter(profile__role="teacher").prefetch_related(
        "courses_taught__lessons", "courses_taught__exams"
    )
    for u in teachers:
        courses = list(u.courses_taught.all())
        total_lessons = sum(c.lessons.count() for c in courses)
        total_exams = sum(c.exams.count() for c in courses)
        ws.append(
            [
                u.get_full_name() or u.username,
                len(courses),
                total_lessons,
                total_exams,
            ]
        )

    _style_header(ws)
    return wb


def build_teacher_performance_report():
    """
    6) Teacher Performance Report
    Columns: Teacher Name, Total Students Across Courses, Average Student Performance (%)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Teacher Performance"
    headers = [
        "Teacher Name",
        "Total Students Across Courses",
        "Average Student Performance (%)",
    ]
    ws.append(headers)

    teachers = (
        User.objects.filter(profile__role="teacher")
        .annotate(
            total_students=Count(
                "courses_taught__enrollments__student", distinct=True
            ),
            avg_percentage=Avg(
                "courses_taught__exams__results__percentage",
                filter=Q(courses_taught__exams__results__is_completed=True),
            ),
        )
        .order_by("username")
    )

    for t in teachers:
        ws.append(
            [
                t.get_full_name() or t.username,
                t.total_students or 0,
                round(t.avg_percentage or 0.0, 2),
            ]
        )

    _style_header(ws)
    return wb


def build_course_teacher_report():
    """
    7) Course-wise Teacher Report
    Columns: Teacher Name, Course Name, Enrollment Count
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Course Teacher"
    headers = ["Teacher Name", "Course Name", "Enrollment Count"]
    ws.append(headers)

    courses = (
        Course.objects.annotate(enrollment_count=Count("enrollments", distinct=True))
        .select_related("teacher")
        .order_by("teacher__username", "title")
    )
    for c in courses:
        ws.append(
            [
                c.teacher.get_full_name() or c.teacher.username,
                c.title,
                c.enrollment_count,
            ]
        )

    _style_header(ws)
    return wb


# -------------------------------------------------------------------
# Course-related reports
# -------------------------------------------------------------------

def build_course_enrollment_report():
    """
    8) Course Enrollment Summary
    Columns: Course Name, Teacher Name, Total Students Enrolled
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Course Enrollment"
    headers = ["Course Name", "Teacher Name", "Total Students Enrolled"]
    ws.append(headers)

    qs = (
        Course.objects.annotate(total_students=Count("enrollments", distinct=True))
        .select_related("teacher")
        .order_by("title")
    )
    for c in qs:
        ws.append(
            [
                c.title,
                c.teacher.get_full_name() or c.teacher.username,
                c.total_students,
            ]
        )

    _style_header(ws)
    return wb


def build_course_completion_report():
    """
    9) Course Completion Report
    Columns: Course Name, Students Completed, Total Enrollments, Completion Rate (%)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Course Completion"
    headers = [
        "Course Name",
        "Students Completed",
        "Total Enrollments",
        "Completion Rate (%)",
    ]
    ws.append(headers)

    courses = Course.objects.annotate(
        total_enrollments=Count("enrollments", distinct=True),
        students_completed=Count(
            "enrollments",
            filter=Q(enrollments__completed=True),
            distinct=True,
        ),
    ).order_by("title")

    for c in courses:
        total = c.total_enrollments or 0
        completed = c.students_completed or 0
        rate = round((completed / total) * 100, 2) if total > 0 else 0.0
        ws.append([c.title, completed, total, rate])

    _style_header(ws)
    return wb


def build_course_revenue_report():
    """
    10) Course Revenue Report (for paid courses)
    Columns: Course Name, Price, Total Enrollments, Total Revenue
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Course Revenue"
    headers = ["Course Name", "Price", "Total Enrollments", "Total Revenue"]
    ws.append(headers)

    courses = (
        Course.objects.filter(price__isnull=False)
        .annotate(total_enrollments=Count("enrollments", distinct=True))
        .order_by("title")
    )
    for c in courses:
        enrollments = c.total_enrollments or 0
        price = c.price or 0
        revenue = price * enrollments
        ws.append([c.title, float(price), enrollments, float(revenue)])

    _style_header(ws)
    return wb


# -------------------------------------------------------------------
# Exam-related reports
# -------------------------------------------------------------------

def build_exam_result_report():
    """
    11) Exam Result Report
    Columns: Exam Name, Course, Student, Marks, Total Marks, Percentage
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Results"
    headers = [
        "Exam Name",
        "Course",
        "Student",
        "Marks",
        "Total Marks",
        "Percentage",
    ]
    ws.append(headers)

    results = (
        ExamResult.objects.filter(is_completed=True)
        .select_related("exam", "exam__course", "student")
        .order_by("exam__title", "student__username")
    )
    for r in results:
        ws.append(
            [
                r.exam.title,
                r.exam.course.title,
                r.student.get_full_name() or r.student.username,
                r.score,
                r.total_marks,
                round(r.percentage, 2),
            ]
        )

    _style_header(ws)
    return wb


def build_exam_pass_fail_report(pass_threshold=50.0):
    """
    12) Pass/Fail Report
    Columns: Exam Name, Course, Total Students, Passed, Failed
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Pass/Fail"
    headers = ["Exam Name", "Course", "Total Students", "Passed", "Failed"]
    ws.append(headers)

    exams = Exam.objects.annotate(
        total_students=Count(
            "results",
            filter=Q(results__is_completed=True),
            distinct=True,
        ),
        passed=Count(
            "results",
            filter=Q(
                results__is_completed=True,
                results__percentage__gte=pass_threshold,
            ),
            distinct=True,
        ),
    ).select_related("course")

    for e in exams:
        total = e.total_students or 0
        passed = e.passed or 0
        failed = total - passed
        ws.append(
            [
                e.title,
                e.course.title,
                total,
                passed,
                failed,
            ]
        )

    _style_header(ws)
    return wb


def build_exam_score_stats_report():
    """
    13) Highest & Lowest Score Report
    Columns: Exam Name, Course, Highest Marks, Lowest Marks, Average Marks
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Score Stats"
    headers = [
        "Exam Name",
        "Course",
        "Highest Marks",
        "Lowest Marks",
        "Average Marks",
    ]
    ws.append(headers)

    exams = Exam.objects.annotate(
        highest=Max("results__score"),
        lowest=Min("results__score"),
        average=Avg(
            "results__score",
            filter=Q(results__is_completed=True),
        ),
    ).select_related("course")

    for e in exams:
        ws.append(
            [
                e.title,
                e.course.title,
                e.highest or 0,
                e.lowest or 0,
                round(e.average or 0.0, 2),
            ]
        )

    _style_header(ws)
    return wb


# -------------------------------------------------------------------
# Platform-level reports
# -------------------------------------------------------------------

def build_platform_summary_report():
    """
    14) Overall System Summary
    Columns: Metric, Count
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Platform Summary"
    ws.append(["Metric", "Count"])

    total_students = UserProfile.objects.filter(role="student").count()
    total_teachers = UserProfile.objects.filter(role="teacher").count()
    total_courses = Course.objects.count()
    total_lessons = Lesson.objects.count()
    total_exams = Exam.objects.count()
    total_enrollments = Enrollment.objects.count()

    rows = [
        ("Total Students", total_students),
        ("Total Teachers", total_teachers),
        ("Total Courses", total_courses),
        ("Total Lessons", total_lessons),
        ("Total Exams", total_exams),
        ("Total Enrollments", total_enrollments),
    ]
    for row in rows:
        ws.append(row)

    _style_header(ws)
    return wb


def build_monthly_growth_report():
    """
    15) Monthly Growth Report
    Columns: Month, New Students, New Courses, New Enrollments
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Growth"
    headers = ["Month", "New Students", "New Courses", "New Enrollments"]
    ws.append(headers)

    # Students
    students = (
        User.objects.filter(profile__role="student")
        .annotate(month=TruncMonth("date_joined"))
        .values("month")
        .annotate(c=Count("id"))
    )
    students_map = {row["month"].date(): row["c"] for row in students if row["month"]}

    # Courses
    courses = (
        Course.objects.annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(c=Count("id"))
    )
    courses_map = {row["month"].date(): row["c"] for row in courses if row["month"]}

    # Enrollments
    enrollments = (
        Enrollment.objects.annotate(month=TruncMonth("enrolled_at"))
        .values("month")
        .annotate(c=Count("id"))
    )
    enrollments_map = {
        row["month"].date(): row["c"] for row in enrollments if row["month"]
    }

    # Union of all months
    all_months = set(students_map) | set(courses_map) | set(enrollments_map)
    for month in sorted(all_months):
        label = month.strftime("%Y-%m")
        ws.append(
            [
                label,
                students_map.get(month, 0),
                courses_map.get(month, 0),
                enrollments_map.get(month, 0),
            ]
        )

    _style_header(ws)
    return wb


def build_most_popular_course_report():
    """
    16) Most Popular Course Report
    Columns: Course Name, Teacher Name, Enrollment Count
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Most Popular Course"
    headers = ["Course Name", "Teacher Name", "Enrollment Count"]
    ws.append(headers)

    course = (
        Course.objects.annotate(enrollment_count=Count("enrollments", distinct=True))
        .select_related("teacher")
        .order_by("-enrollment_count")
        .first()
    )

    if course:
        ws.append(
            [
                course.title,
                course.teacher.get_full_name() or course.teacher.username,
                course.enrollment_count or 0,
            ]
        )

    _style_header(ws)
    return wb


def build_most_active_teacher_report():
    """
    17) Most Active Teacher Report
    Columns: Teacher Name, Total Courses, Total Lessons, Total Exams, Activity Score
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Most Active Teacher"
    headers = [
        "Teacher Name",
        "Total Courses",
        "Total Lessons",
        "Total Exams",
        "Activity Score",
    ]
    ws.append(headers)

    teachers = (
        User.objects.filter(profile__role="teacher")
        .annotate(
            total_courses=Count("courses_taught", distinct=True),
            total_lessons=Count("courses_taught__lessons", distinct=True),
            total_exams=Count("courses_taught__exams", distinct=True),
        )
        .annotate(
            activity_score=F("total_courses") + F("total_lessons") + F("total_exams")
        )
    )

    most_active = teachers.order_by("-activity_score").first()
    if most_active:
        ws.append(
            [
                most_active.get_full_name() or most_active.username,
                most_active.total_courses or 0,
                most_active.total_lessons or 0,
                most_active.total_exams or 0,
                most_active.activity_score or 0,
            ]
        )

    _style_header(ws)
    return wb


def save_workbook_to_buffer(wb):
    """Write workbook to in-memory buffer and return bytes."""
    from io import BytesIO

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
