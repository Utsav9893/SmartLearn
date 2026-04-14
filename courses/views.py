from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Avg
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.core.exceptions import ValidationError
from django.utils import timezone
from .models import Course, Lesson, Enrollment, LessonCompletion, Feedback
from .forms import FeedbackForm
from accounts.decorators import teacher_required, student_required, teacher_or_admin_required
from exams.models import Exam, ExamResult
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import datetime


@login_required
@teacher_required
def course_list(request):
    """
    List all courses created by the teacher
    """
    courses = Course.objects.filter(teacher=request.user)
    return render(request, 'courses/course_list.html', {
        'courses': courses,
        'back_url': reverse('accounts:teacher_dashboard'),
    })


@login_required
@teacher_required
def course_create(request):
    """
    Create a new course
    """
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        is_published = request.POST.get('is_published') == 'on'
        
        course = Course.objects.create(
            title=title,
            description=description,
            teacher=request.user,
            is_published=is_published
        )
        messages.success(request, f'Course "{course.title}" created successfully!')
        return redirect('courses:course_detail', pk=course.pk)
    
    return render(request, 'courses/course_form.html', {'action': 'Create'})


@login_required
@teacher_or_admin_required
def course_detail(request, pk):
    """
    Course detail view - shows lessons and allows editing
    """
    course = get_object_or_404(Course, pk=pk)
    
    # Check if user has permission (teacher or admin)
    if not (course.teacher == request.user or request.user.profile.is_admin()):
        messages.error(request, 'You do not have permission to view this course.')
        return redirect('accounts:dashboard')
    
    lessons = course.lessons.all()
    exams = Exam.objects.filter(course=course)
    
    # Summarized numeric data for teacher/admin
    total_enrolled = Enrollment.objects.filter(course=course).count()
    exam_attempts = ExamResult.objects.filter(exam__course=course, is_completed=True).values('student').distinct().count()
    
    context = {
        'course': course,
        'lessons': lessons,
        'exams': exams,
        'total_enrolled': total_enrolled,
        'exam_attempts': exam_attempts,
        'back_url': reverse('accounts:teacher_dashboard'),
    }
    return render(request, 'courses/course_detail.html', context)


@login_required
@teacher_required
def course_report_excel(request, pk):
    """
    Generate Excel report for a specific course
    """
    course = get_object_or_404(Course, pk=pk, teacher=request.user)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Course Report"
    
    headers = ["Student Name", "Enrollment Status", "Exam Attempt Status", "Marks Obtained"]
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    enrollments = Enrollment.objects.filter(course=course).select_related('student')
    
    for enrollment in enrollments:
        student = enrollment.student
        enrollment_status = "Completed" if enrollment.completed else "In Progress"
        
        # Check for any exam results in this course
        results = ExamResult.objects.filter(exam__course=course, student=student, is_completed=True)
        if results.exists():
            attempt_status = "Attempted"
            # Get highest score or list all? Let's show the best score among all exams in the course
            best_result = results.order_by('-percentage').first()
            marks_info = f"{best_result.score}/{best_result.total_marks} ({best_result.percentage}%)"
        else:
            attempt_status = "Not Attempted"
            marks_info = "N/A"
            
        ws.append([
            student.get_full_name() or student.username,
            enrollment_status,
            attempt_status,
            marks_info
        ])
        
    # Auto-adjust column widths
    for col in range(1, ws.max_column + 1):
        max_len = 0
        column = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            try:
                if len(str(ws[f"{column}{row}"].value)) > max_len:
                    max_len = len(str(ws[f"{column}{row}"].value))
            except:
                pass
        ws.column_dimensions[column].width = max_len + 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{course.title}_Report_{datetime.date.today()}.xlsx"'
    wb.save(response)
    return response


@login_required
@teacher_required
def course_edit(request, pk):
    """
    Edit course
    """
    course = get_object_or_404(Course, pk=pk, teacher=request.user)
    
    if request.method == 'POST':
        course.title = request.POST.get('title')
        course.description = request.POST.get('description')
        course.is_published = request.POST.get('is_published') == 'on'
        course.save()
        messages.success(request, 'Course updated successfully!')
        return redirect('courses:course_detail', pk=course.pk)
    
    return render(request, 'courses/course_form.html', {'course': course, 'action': 'Edit'})


@login_required
@teacher_required
def course_delete(request, pk):
    """
    Delete course
    """
    course = get_object_or_404(Course, pk=pk, teacher=request.user)
    
    if request.method == 'POST':
        course_title = course.title
        course.delete()
        messages.success(request, f'Course "{course_title}" deleted successfully!')
        return redirect('courses:course_list')
    
    return render(request, 'courses/course_confirm_delete.html', {'course': course})


@login_required
@teacher_required
def lesson_create(request, course_pk):
    """
    Create a new lesson for a course
    """
    course = get_object_or_404(Course, pk=course_pk, teacher=request.user)
    
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        lesson_type = request.POST.get('lesson_type')
        
        # Calculate lesson order automatically
        last_lesson = course.lessons.order_by('-order').first()
        order = (last_lesson.order + 1) if last_lesson else 1
        
        video_file = request.FILES.get('video_file')
        pdf_file = request.FILES.get('pdf_file')
        
        try:
            lesson = Lesson(
                course=course,
                title=title,
                description=description,
                lesson_type=lesson_type,
                order=order,
                video_file=video_file,
                pdf_file=pdf_file
            )
            lesson.save()
            messages.success(request, 'Lesson created successfully!')
            return redirect('courses:course_detail', pk=course.pk)
        except ValidationError as e:
            for field, errors in e.message_dict.items():
                for error in errors:
                    messages.error(request, error)
        except Exception as e:
            messages.error(request, f'Error creating lesson: {str(e)}')
    
    return render(request, 'courses/lesson_form.html', {'course': course, 'action': 'Create'})


@login_required
@teacher_required
def lesson_edit(request, pk):
    """
    Edit lesson
    """
    lesson = get_object_or_404(Lesson, pk=pk)
    
    if lesson.course.teacher != request.user:
        messages.error(request, 'You do not have permission to edit this lesson.')
        return redirect('accounts:dashboard')
    
    if request.method == 'POST':
        lesson.title = request.POST.get('title')
        lesson.description = request.POST.get('description')
        lesson.lesson_type = request.POST.get('lesson_type')
        
        if request.FILES.get('video_file'):
            lesson.video_file = request.FILES.get('video_file')
        if request.FILES.get('pdf_file'):
            lesson.pdf_file = request.FILES.get('pdf_file')
        
        try:
            lesson.save()
            messages.success(request, 'Lesson updated successfully!')
            return redirect('courses:course_detail', pk=lesson.course.pk)
        except ValidationError as e:
            for field, errors in e.message_dict.items():
                for error in errors:
                    messages.error(request, error)
        except Exception as e:
            messages.error(request, f'Error updating lesson: {str(e)}')
    
    return render(request, 'courses/lesson_form.html', {'lesson': lesson, 'action': 'Edit'})


@login_required
@teacher_required
def lesson_delete(request, pk):
    """
    Delete lesson
    """
    lesson = get_object_or_404(Lesson, pk=pk)
    
    if lesson.course.teacher != request.user:
        messages.error(request, 'You do not have permission to delete this lesson.')
        return redirect('accounts:dashboard')
    
    course_pk = lesson.course.pk
    if request.method == 'POST':
        lesson.delete()
        messages.success(request, 'Lesson deleted successfully!')
        return redirect('courses:course_detail', pk=course_pk)
    
    return render(request, 'courses/lesson_confirm_delete.html', {'lesson': lesson})


@login_required
@student_required
def course_browse(request):
    """
    Browse all available courses for students
    """
    query = request.GET.get('q')
    if query:
        courses = Course.objects.filter(title__icontains=query, is_published=True)
    else:
        courses = Course.objects.filter(is_published=True)
        
    enrolled_course_ids = request.user.courses_enrolled.values_list('id', flat=True)
    
    context = {
        'courses': courses,
        'enrolled_course_ids': enrolled_course_ids,
        'query': query,
    }
    return render(request, 'courses/course_browse.html', context)


@login_required
@student_required
def course_detail_browse(request, pk):
    """
    Course detail for students: limited info when not enrolled (no lessons);
    when enrolled, redirect to course view or show Go to Course.
    """
    course = get_object_or_404(Course, pk=pk, is_published=True)
    is_enrolled = Enrollment.objects.filter(student=request.user, course=course).exists()

    if is_enrolled:
        return redirect('courses:course_view', pk=course.pk)

    context = {
        'course': course,
        'back_url': reverse('courses:course_browse'),
    }
    return render(request, 'courses/course_detail_browse.html', context)


@login_required
@student_required
def course_enroll(request, pk):
    """
    Enroll in a course
    """
    course = get_object_or_404(Course, pk=pk, is_published=True)
    
    if course.students.filter(id=request.user.id).exists():
        messages.info(request, 'You are already enrolled in this course.')
    else:
        course.students.add(request.user)
        Enrollment.objects.get_or_create(student=request.user, course=course)
        messages.success(request, f'Successfully enrolled in "{course.title}"!')
    
    return redirect('courses:course_view', pk=course.pk)


@login_required
@student_required
def course_view(request, pk):
    """
    Student view of enrolled course
    """
    course = get_object_or_404(Course, pk=pk)
    
    # Backend enrollment check using Enrollment model
    if not Enrollment.objects.filter(student=request.user, course=course).exists():
        messages.error(request, 'You are not enrolled in this course.')
        return redirect('courses:course_browse')
    
    lessons = course.lessons.all().order_by('order')
    exams = Exam.objects.filter(course=course, is_published=True)
    
    # Exam results for this student (exam_id -> result pk) for exams already completed
    completed_exam_result_pks = {}
    for er in ExamResult.objects.filter(
        student=request.user,
        exam__in=exams,
        is_completed=True
    ):
        completed_exam_result_pks[er.exam_id] = er.pk
    
    # List of (exam, result_pk or None) for template
    exams_with_result = [(exam, completed_exam_result_pks.get(exam.id)) for exam in exams]
    
    # Calculate progress
    enrollment = Enrollment.objects.filter(student=request.user, course=course).first()
    progress_percentage = enrollment.get_progress_percentage() if enrollment else 0
    completed_results = ExamResult.objects.filter(
        student=request.user,
        exam__course=course,
        is_completed=True
    )
    average_score = completed_results.aggregate(avg_percentage=Avg('percentage')).get('avg_percentage') or 0
    feedback_obj = Feedback.objects.filter(student=request.user, course=course).first()
    feedback_exists = feedback_obj is not None
    can_download_certificate = (
        progress_percentage >= 100 and
        average_score >= 80 and
        feedback_exists and
        completed_results.exists()
    )

    if average_score < 40:
        recommendation = "You should focus on beginner-level courses to strengthen your basics."
    elif average_score < 70:
        recommendation = "You are doing well. Try intermediate-level challenges to improve further."
    else:
        recommendation = "Excellent performance! You are ready for advanced-level courses."

    # Feedback is allowed only once, and only when course progress is complete.
    feedback_form = FeedbackForm()
    
    # Get completed lesson IDs for highlighting
    completed_lesson_ids = LessonCompletion.objects.filter(
        student=request.user,
        lesson__course=course
    ).values_list('lesson_id', flat=True)
    
    context = {
        'course': course,
        'lessons': lessons,
        'exams': exams,
        'exams_with_result': exams_with_result,
        'progress_percentage': progress_percentage,
        'average_score': round(average_score, 2),
        'recommendation': recommendation,
        'eligible': can_download_certificate,
        'can_download_certificate': can_download_certificate,
        'feedback_form': feedback_form,
        'feedback_exists': feedback_exists,
        'feedback_obj': feedback_obj,
        'completed_lesson_ids': completed_lesson_ids,
        'back_url': reverse('accounts:student_dashboard'),
    }
    return render(request, 'courses/course_view.html', context)


@login_required
@student_required
def submit_feedback(request, course_id):
    """
    Submit feedback for a course once the student reaches 100% progress.
    """
    if request.method != 'POST':
        return redirect('courses:course_view', pk=course_id)

    course = get_object_or_404(Course, pk=course_id)
    enrollment = get_object_or_404(Enrollment, student=request.user, course=course)

    if enrollment.get_progress_percentage() < 100:
        messages.error(request, 'You can submit feedback only after completing 100% of the course.')
        return redirect('courses:course_view', pk=course.pk)

    if Feedback.objects.filter(student=request.user, course=course).exists():
        messages.info(request, 'Feedback already submitted for this course.')
        return redirect('courses:course_view', pk=course.pk)

    form = FeedbackForm(request.POST)
    if form.is_valid():
        feedback = form.save(commit=False)
        feedback.student = request.user
        feedback.course = course
        feedback.save()
        messages.success(request, 'Thank you! Your feedback has been submitted successfully.')
    else:
        messages.error(request, 'Please provide valid feedback details.')

    return redirect('courses:course_view', pk=course.pk)


@login_required
@student_required
def generate_certificate(request, course_id):
    """
    Generate certificate PDF for eligible students only.
    Eligibility:
    - Course progress must be 100%
    - Average completed exam score in course must be >= 80%
    - Feedback must be submitted
    """
    course = get_object_or_404(Course, pk=course_id)
    enrollment = get_object_or_404(Enrollment, student=request.user, course=course)

    progress_percentage = enrollment.get_progress_percentage()
    completed_results = ExamResult.objects.filter(
        student=request.user,
        exam__course=course,
        is_completed=True
    )
    average_score = completed_results.aggregate(avg_percentage=Avg('percentage')).get('avg_percentage') or 0
    feedback_exists = Feedback.objects.filter(student=request.user, course=course).exists()

    if progress_percentage < 100 or average_score < 80 or not feedback_exists or not completed_results.exists():
        messages.error(
            request,
            'Certificate unlocks after 100% progress, average score of at least 80%, and feedback submission.'
        )
        return redirect('courses:course_view', pk=course.pk)

    completion_date = enrollment.completed_at or timezone.now()
    display_name = request.user.get_full_name() or request.user.username

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="certificate_course_{course.pk}_{request.user.pk}.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    pdf.setTitle('Course Completion Certificate')
    pdf.setFont('Helvetica-Bold', 28)
    pdf.drawCentredString(width / 2, height - 120, 'Certificate of Completion')

    pdf.setFont('Helvetica', 14)
    pdf.drawCentredString(width / 2, height - 190, 'This is to certify that')

    pdf.setFont('Helvetica-Bold', 22)
    pdf.drawCentredString(width / 2, height - 240, display_name)

    pdf.setFont('Helvetica', 14)
    pdf.drawCentredString(width / 2, height - 290, 'has successfully completed the course')

    pdf.setFont('Helvetica-Bold', 18)
    pdf.drawCentredString(width / 2, height - 330, course.title)

    pdf.setFont('Helvetica', 12)
    pdf.drawCentredString(
        width / 2,
        height - 380,
        f'Completion Date: {completion_date.strftime("%d %B %Y")}'
    )
    pdf.drawCentredString(
        width / 2,
        height - 405,
        f'Average Score: {round(average_score, 2)}%'
    )

    pdf.showPage()
    pdf.save()
    return response


@login_required
@student_required
def lesson_view(request, pk):
    """
    Student view of a lesson
    """
    lesson = get_object_or_404(Lesson, pk=pk)
    course = lesson.course
    
    # Backend enrollment check using Enrollment model
    if not Enrollment.objects.filter(student=request.user, course=course).exists():
        messages.error(request, 'You are not enrolled in this course.')
        return redirect('courses:course_browse')
    
    # Mark lesson as completed when student views it
    LessonCompletion.objects.get_or_create(
        student=request.user,
        lesson=lesson
    )
    
    # Get next and previous lessons
    lessons = list(course.lessons.all().order_by('order'))
    current_index = lessons.index(lesson) if lesson in lessons else -1
    next_lesson = lessons[current_index + 1] if current_index >= 0 and current_index < len(lessons) - 1 else None
    prev_lesson = lessons[current_index - 1] if current_index > 0 else None
    
    # Calculate progress for display
    enrollment = Enrollment.objects.filter(student=request.user, course=course).first()
    progress_percentage = enrollment.get_progress_percentage() if enrollment else 0
    
    context = {
        'lesson': lesson,
        'course': course,
        'next_lesson': next_lesson,
        'prev_lesson': prev_lesson,
        'progress_percentage': progress_percentage,
        'back_url': reverse('courses:course_view', args=[course.pk]),
    }
    return render(request, 'courses/lesson_view.html', context)
